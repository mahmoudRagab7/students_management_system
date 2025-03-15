import os
import json
try:
    import openpyxl
    excel_available = True
except ImportError:
    excel_available = False

class DatabaseHandler:
    """Handles file operations for student data."""
    
    def __init__(self, filename="students.txt"):
        """Initialize with the database filename."""
        self.filename = filename
    
    def load_students(self):
        """Load students from the file."""
        from models import Student
        
        students = []
        
        # Check if file exists
        if not os.path.exists(self.filename):
            return students
        
        try:
            with open(self.filename, 'r') as file:
                data = file.read().strip()
                if not data:
                    return students
                
                # Parse the JSON data
                students_data = json.loads(data)
                
                # Create Student objects
                for student_dict in students_data:
                    students.append(Student.from_dict(student_dict))
                
                return students
        except (json.JSONDecodeError, FileNotFoundError) as e:
            print(f"Error loading students: {e}")
            return students
    
    def save_students(self, students):
        """Save students to the file."""
        try:
            # Convert Student objects to dictionaries
            students_data = [student.to_dict() for student in students]
            
            # Write to file as JSON
            with open(self.filename, 'w') as file:
                json.dump(students_data, file, indent=2)
            
            return True
        except Exception as e:
            print(f"Error saving students: {e}")
            return False
    
    def export_to_csv(self, students, filename="students.csv"):
        """Export student data to CSV format."""
        try:
            # Get absolute path for the file
            abs_path = os.path.abspath(filename)
            
            with open(filename, 'w') as file:
                # Write header
                file.write("Student ID,Name,Age,Gender,Courses\n")
                
                # Write each student's data
                for student in students:
                    courses = "|".join(student.courses)
                    file.write(f"{student.student_id},{student.name},{student.age},{student.gender},{courses}\n")
            
            return f"Student data exported successfully to {filename}\nFile location: {abs_path}"
        except Exception as e:
            return f"Error exporting to CSV: {e}"
    
    def export_to_excel(self, students, filename="students.xlsx"):
        """Export student data to Excel format."""
        if not excel_available:
            return "Excel export is not available. Please install openpyxl with 'pip install openpyxl'"
        
        try:
            # Get absolute path for the file
            abs_path = os.path.abspath(filename)
            
            # Create a new workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Students"
            
            # Add headers
            headers = ["Student ID", "Name", "Age", "Gender", "Courses"]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                # Optional: Style the header
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Add student data
            for row_num, student in enumerate(students, 2):
                worksheet.cell(row=row_num, column=1).value = student.student_id
                worksheet.cell(row=row_num, column=2).value = student.name
                worksheet.cell(row=row_num, column=3).value = student.age
                worksheet.cell(row=row_num, column=4).value = student.gender
                worksheet.cell(row=row_num, column=5).value = ", ".join(student.courses)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            workbook.save(filename)
            return f"Student data exported successfully to {filename}\nFile location: {abs_path}"
        except Exception as e:
            return f"Error exporting to Excel: {e}"